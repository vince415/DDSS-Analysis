import streamlit as st
import pandas as pd
import openpyxl
import io
import hashlib
from datetime import datetime, timedelta
import plotly.graph_objects as go
import plotly.express as px
import re
from calendar import monthrange

st.set_page_config(
    page_title="DDSS Forecast Analysis",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ═══════════════════════════════════════════════════════════════
#  Utility Functions
# ═══════════════════════════════════════════════════════════════

def normalize(s):
    """Normalize string: strip + uppercase"""
    return str(s).strip().upper() if s is not None else ""


def extract_mpa_year_week(filename):
    """
    Extract MPA, year, and week from filename

    Supports formats:
    - "FXN 4X 2026wk9.xlsx"      → MPA="FXN 4X", Year=2026, Week=9
    - "FXN 2026 wk 9.xlsx"       → MPA="FXN", Year=2026, Week=9
    - "TH 2026wk10.xlsx"         → MPA="TH", Year=2026, Week=10
    - "FXN_2026wk3.xlsx"         → MPA="FXN", Year=2026, Week=3
    - "FXN wk9.xlsx"             → MPA="FXN", Year=None, Week=9
    """
    stem = re.sub(r'\.xlsx?$', '', filename, flags=re.IGNORECASE).strip()

    # Pattern 1: "Name YYYYwkNN" or "Name YYYY wk NN"
    # Handles spaces in name like "FXN 4X 2026wk9"
    m = re.search(r'^(.+?)\s+(\d{4})\s*wk\s*(\d+)\s*$', stem, re.IGNORECASE)
    if m:
        mpa = m.group(1).strip().upper() if m.group(1).strip() else None
        year = int(m.group(2))
        week = int(m.group(3))
        return mpa, year, week

    # Pattern 2: "Name_YYYYwkNN" or "Name-YYYYwkNN"
    m = re.search(r'^(.+?)[_\-]\s*(\d{4})\s*wk\s*(\d+)\s*$', stem, re.IGNORECASE)
    if m:
        mpa = m.group(1).strip().upper() if m.group(1).strip() else None
        year = int(m.group(2))
        week = int(m.group(3))
        return mpa, year, week

    # Pattern 3: "Name wkNN" (no year)
    m = re.search(r'^(.+?)\s+wk\s*(\d+)\s*$', stem, re.IGNORECASE)
    if m:
        mpa = m.group(1).strip().upper() if m.group(1).strip() else None
        week = int(m.group(2))
        return mpa, None, week

    return None, None, None


def create_week_sort_key(year, week):
    """Create sortable key from year and week"""
    return year * 100 + week if year is not None else week


def find_all_ddss_sheets(wb):
    """Find all DDSS-like sheets (case-insensitive)"""
    results = []
    for name in wb.sheetnames:
        norm_name = name.strip().upper()
        if norm_name == 'DDSS' or norm_name.startswith("DDSS W"):
            results.append((name, wb[name]))
    return results


def detect_header_cols(row):
    """
    Detect column indices from header row (case-insensitive + aliases)

    Supports multiple column name variations:
    - Type / Details / Detail
    - Consign PN / Part Number / PN / Part No / Partnumber
    - Data Description / Description / Desc
    - On hand (RM) / On hand (FG) / On hand / Onhand
    """
    norm = [normalize(c) for c in row]

    def find(*candidates):
        for c in candidates:
            cn = normalize(c)
            for i, v in enumerate(norm):
                if v == cn:
                    return i
        return None

    # MPA column
    col_mpa = find('MPA')

    # Type column - supports multiple aliases
    col_type = find('TYPE', 'DETAILS', 'DETAIL')

    # Part Number column - supports multiple aliases
    col_part = find(
        'CONSIGN PN',
        'PART NUMBER',
        'PN',
        'PARTNUMBER',
        'PART NO',
        'PART NO.',
        'PARTNO',
        'CONSIGN'
    )

    # Data Description column - supports multiple aliases
    col_desc = find(
        'DATA DESCRIPTION',
        'DESCRIPTION',
        'DESC'
    )

    # On hand column - supports multiple aliases
    col_oh = find(
        'ON HAND (RM)',
        'ON HAND (FG)',
        'ON HAND',
        'ONHAND'
    )

    # Required columns check
    if col_mpa is None or col_part is None or col_desc is None:
        return None

    # Find first date column
    col_date = next((i for i, v in enumerate(row) if isinstance(v, datetime)), None)
    if col_date is None:
        return None

    return {
        'mpa': col_mpa, 'type': col_type, 'part': col_part,
        'desc': col_desc, 'oh': col_oh, 'date0': col_date
    }


# ═══════════════════════════════════════════════════════════════
#  SDOS Loading and WOS Calculation
# ═══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def load_sdos_file(sdos_bytes: bytes, mpa_from_filename: str = None):
    """
    Load SDOS file and return {Product_ID: {date: sdos_value}}

    Location ID to MPA mapping:
    - SG5HVC → FXN (includes 2X and 4X)
    - 01EMVL → TH
    - 02AMVC → VN

    Filters by Location ID based on MPA from filename
    """
    try:
        # MPA to Location mapping
        mpa_to_location = {
            'FXN': 'SG5HVC',
            'TH': '01EMVL',
            'VN': '02AMVC'
        }

        # Determine target location from MPA
        target_location = None
        if mpa_from_filename:
            mpa_upper = mpa_from_filename.strip().upper()
            for mpa_key, location_id in mpa_to_location.items():
                if mpa_key in mpa_upper:
                    target_location = location_id
                    break

        xl = pd.ExcelFile(io.BytesIO(sdos_bytes))

        if 'SDOS' not in xl.sheet_names:
            st.warning("SDOS sheet not found")
            return {}

        raw = xl.parse('SDOS', header=None)

        # Find header row (usually in rows 0-4)
        header_row_idx = None
        for idx in range(0, 5):
            if idx >= len(raw):
                break
            row = raw.iloc[idx]
            row_str = [str(v).strip().upper() if pd.notna(v) else "" for v in row]

            # Look for key column names
            if any(name in row_str for name in ['PRODUCT ID', 'PART NUMBER', 'CONSIGN PN', 'PN']):
                header_row_idx = idx
                break

        if header_row_idx is None:
            st.error("SDOS header row not found")
            return {}

        header = raw.iloc[header_row_idx]
        header_norm = [str(v).strip().upper() if pd.notna(v) else "" for v in header]

        # Find required columns
        location_col = next((i for i, n in enumerate(header_norm)
                             if n in ['LOCATION ID', 'LOCATION', 'LOC ID']), None)
        product_col = next((i for i, n in enumerate(header_norm)
                            if n in ['PRODUCT ID', 'PART NUMBER', 'CONSIGN PN', 'PN', 'PARTNUMBER', 'PART NO']), None)
        keyfigure_col = next((i for i, n in enumerate(header_norm)
                              if n in ['KEYFIGURE', 'KEY FIGURE', 'KEY_FIGURE']), None)

        if product_col is None or keyfigure_col is None:
            st.error("Required columns not found: Product ID or KeyFigure")
            return {}

        # Find date columns (year >= 2020)
        date_columns = {i: pd.Timestamp(v) for i, v in enumerate(header)
                        if pd.notna(v) and hasattr(v, 'year') and v.year >= 2020}

        if not date_columns:
            st.error("Date columns not found")
            return {}

        # Build SDOS data dictionary
        sdos_data = {}
        filtered_count = 0

        for row_idx in range(header_row_idx + 1, len(raw)):
            location_id = raw.iloc[row_idx, location_col] if location_col is not None else None
            product_id = raw.iloc[row_idx, product_col]
            keyfigure = raw.iloc[row_idx, keyfigure_col]

            # Skip if not Safety Days of Supply
            if pd.isna(product_id) or str(keyfigure).strip() != 'Safety Days of Supply':
                continue

            # Filter by target location if specified
            if target_location and location_col is not None:
                location_str = str(location_id).strip().upper() if pd.notna(location_id) else ""
                if location_str != target_location:
                    filtered_count += 1
                    continue

            product_id_str = str(product_id).strip()
            if product_id_str not in sdos_data:
                sdos_data[product_id_str] = {}

            # Extract SDOS values for all dates
            for col_idx, date in date_columns.items():
                sdos_val = raw.iloc[row_idx, col_idx]
                if pd.notna(sdos_val) and sdos_val > 0:
                    sdos_data[product_id_str][date] = float(sdos_val)

        # Success message
        if sdos_data:
            msg = f"✓ SDOS loaded: {len(sdos_data)} parts"
            if target_location:
                msg += f" (Location: {target_location})"
            if filtered_count > 0:
                msg += f" (filtered {filtered_count} other locations)"
            st.success(msg)
        else:
            st.warning("⚠️ No valid SDOS data found")

        return sdos_data

    except Exception as e:
        st.error(f"SDOS loading error: {e}")
        return {}


def get_quarter_end_month(data_date):
    """Get quarter-end month: Jan/Apr/Jul/Oct"""
    month = data_date.month
    if month <= 1:
        return 1
    elif month <= 4:
        return 4
    elif month <= 7:
        return 7
    else:
        return 10


def get_sdos_for_part(product_id, data_date, sdos_data):
    """
    Get SDOS value for part at quarter-end

    Logic:
    1. Find quarter-end month for data_date
    2. Look for SDOS value in that quarter-end month
    3. If not found, try next quarter-end
    4. Fallback to earliest SDOS if no quarter-end data found
    """
    if product_id not in sdos_data:
        return None

    part_sdos = sdos_data[product_id]
    if not part_sdos:
        return None

    quarter_end_month = get_quarter_end_month(data_date)
    year = data_date.year
    quarter_end_months = [1, 4, 7, 10]
    start_idx = quarter_end_months.index(quarter_end_month)

    # Try current and next 3 quarters
    for i in range(4):
        try_month = quarter_end_months[(start_idx + i) % 4]
        try_year = year + (start_idx + i) // 4
        last_day = monthrange(try_year, try_month)[1]
        target_date_end = datetime(try_year, try_month, last_day)

        # Find closest date in this quarter-end month
        best_date = None
        best_diff = None

        for sdos_date, sdos_val in part_sdos.items():
            if sdos_date.year == try_year and sdos_date.month == try_month:
                diff = abs((target_date_end - sdos_date).days)
                if best_date is None or diff < best_diff:
                    best_date = sdos_date
                    best_diff = diff

        if best_date:
            return part_sdos[best_date]

    # Fallback: return earliest SDOS
    if part_sdos:
        earliest_date = min(part_sdos.keys())
        return part_sdos[earliest_date]

    return None


def is_fxn_2x(mpa, type_str):
    """
    Check if FXN 2X (only FXN 2X uses n = SDOS/7 without +1)

    Must contain both 'FXN' and '2X' as standalone words
    """
    if not mpa and not type_str:
        return False

    check_str = f"{mpa or ''} {type_str or ''}".upper()

    if 'FXN' not in check_str:
        return False

    # Use regex to ensure '2X' is a standalone word (not part of '4X')
    return bool(re.search(r'\b2X\b', check_str))


def calculate_wos_for_dataframe(df, sdos_data):
    """
    Calculate WOS for each Part Number

    FIXED: Calculate WOS separately for each week
    Previous bug: All weeks were merged, causing identical WOS values

    Formula: WOS = (Balance / Future_Demand) × n
    where:
    - n = SDOS/7 (for FXN 2X only)
    - n = SDOS/7 + 1 (for all others)
    - Future_Demand = sum of next n weeks' demand components
    """
    if not sdos_data or df.empty:
        return df

    # Demand components to sum
    demand_components = [
        'POR demand',
        'PO vs POR adustment',  # Note: original typo in data
        'Backlog',
        'Build and Hold',
        'Pre-build',
        "Test Req't"
    ]

    # Remove existing WOS rows
    df = df[df['Data_Description'] != 'WOS'].copy()
    wos_rows = []

    # CRITICAL FIX: Group by Week FIRST, then by Part
    # This ensures each week's data is calculated separately
    week_groups = df.groupby(['WeekSortKey', 'WeekLabel', 'Year', 'Filename'])

    for (week_key, week_label, year, filename), week_df in week_groups:
        # Now group by Part + MPA + Sheet + Type within this week
        part_groups = week_df.groupby(['Consign_PN', 'MPA', 'Sheet', 'Type'])

        for (part, mpa, sheet, type_val), part_df in part_groups:
            # Get SDOS value
            first_date = part_df['Date'].min()
            sdos = get_sdos_for_part(part, first_date, sdos_data)
            if sdos is None:
                continue

            # Calculate n weeks
            is_fxn = is_fxn_2x(mpa, type_val)
            n = int(sdos / 7) if is_fxn else int(sdos / 7) + 1

            if n <= 0:
                continue

            # Create pivot for THIS WEEK ONLY
            week_pivot = part_df.pivot_table(
                index='Date',
                columns='Data_Description',
                values='Value',
                aggfunc='first'  # Use 'first' since each date should appear once per week
            ).sort_index()

            if 'Balance' not in week_pivot.columns:
                continue

            # Check for demand components
            available_demand_cols = [c for c in demand_components if c in week_pivot.columns]
            if not available_demand_cols:
                continue

            # Calculate total demand for each date
            week_pivot['TotalDemand'] = week_pivot[available_demand_cols].sum(axis=1).fillna(0)
            week_dates_list = week_pivot.index.tolist()

            # Get forecast dates only (exclude On hand)
            forecast_only = part_df[part_df['Column_Type'] == 'Forecast']
            forecast_dates = sorted(forecast_only['Date'].unique())

            # Calculate WOS for each forecast date in this week
            for date in forecast_dates:
                if date not in week_pivot.index:
                    continue

                balance = week_pivot.loc[date, 'Balance']
                if pd.isna(balance) or balance == 0:
                    continue

                # Find date position
                date_idx = week_dates_list.index(date)

                # Calculate future n weeks demand
                if date_idx + n < len(week_dates_list):
                    future_slice = week_pivot.iloc[date_idx + 1:date_idx + 1 + n]
                    future_demand = future_slice['TotalDemand'].sum()

                    if future_demand > 0:
                        wos_value = (balance / future_demand) * n

                        wos_rows.append({
                            'Week': week_key % 100 if week_key >= 100 else week_key,
                            'Year': year,
                            'WeekSortKey': week_key,
                            'WeekLabel': week_label,
                            'MPA': mpa,
                            'Sheet': sheet,
                            'Filename': filename,
                            'Type': type_val,
                            'Consign_PN': part,
                            'Data_Description': 'WOS',
                            'Date': date,
                            'Column_Type': 'Forecast',
                            'Value': wos_value
                        })

    # Add WOS rows to dataframe
    if wos_rows:
        wos_df = pd.DataFrame(wos_rows)
        df = pd.concat([df, wos_df], ignore_index=True)

    return df


# ═══════════════════════════════════════════════════════════════
#  File Parsing
# ═══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def parse_one_file(file_bytes: bytes, filename: str):
    """
    Parse single Excel file, extract data from all DDSS sheets

    Handles files with repeating headers (grouped by Part Number)
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ddss_sheets = find_all_ddss_sheets(wb)
        if not ddss_sheets:
            return None

        mpa_file, year, week = extract_mpa_year_week(filename)
        if week is None:
            return None

        week_sort_key = create_week_sort_key(year, week)
        all_records = []

        for sheet_name, sheet in ddss_sheets:
            records = []
            cur_part = cur_type = cols = start_date = None

            for row in sheet.iter_rows(values_only=True):
                if not any(row):
                    continue

                # Try to detect header row
                det = detect_header_cols(row)
                if det is not None:
                    cols = det
                    sd = row[cols['date0']]
                    start_date = sd if isinstance(sd, datetime) else None
                    if start_date is None:
                        cols = None
                    # Reset current part and type when header found
                    cur_part = cur_type = None
                    continue

                if not cols or not start_date:
                    continue

                L = len(row)

                # Update current part (skip header column names)
                if cols['part'] < L:
                    rp = row[cols['part']]
                    rp_norm = normalize(rp)
                    # Exclude all possible header column names
                    if rp and rp_norm not in ('CONSIGN PN', 'PART NUMBER', 'PN', 'PARTNUMBER',
                                              'PART NO', 'PART NO.', 'PARTNO', 'CONSIGN', ''):
                        cur_part = str(rp).strip()

                # Update current type (skip header column names)
                if cols['type'] is not None and cols['type'] < L:
                    rt = row[cols['type']]
                    rt_norm = normalize(rt)
                    if rt and rt_norm not in ('TYPE', 'DETAILS', 'DETAIL', ''):
                        cur_type = str(rt).strip()

                # Check Data Description
                if cols['desc'] >= L:
                    continue

                rd = row[cols['desc']]
                rd_norm = normalize(rd)
                # Skip if no part, no description, or description is a header name
                if not cur_part or not rd or rd_norm in ('DATA DESCRIPTION', 'DESCRIPTION', 'DESC', ''):
                    continue

                # Get MPA (from file or from row)
                mpa = mpa_file or (normalize(row[cols['mpa']])
                                   if cols['mpa'] < L and row[cols['mpa']] else None)

                # Create week label
                week_label = f"{year}Wk{week}" if year is not None else f"Wk{week}"

                # Base record
                base = {
                    'Week': week,
                    'Year': year,
                    'WeekSortKey': week_sort_key,
                    'WeekLabel': week_label,
                    'MPA': mpa,
                    'Sheet': sheet_name,
                    'Filename': filename,
                    'Type': cur_type,
                    'Consign_PN': cur_part,
                    'Data_Description': str(rd).strip(),
                }

                # Extract On hand value
                if cols['oh'] is not None and cols['oh'] < L:
                    oh = row[cols['oh']]
                    if isinstance(oh, (int, float)):
                        records.append({
                            **base,
                            'Date': start_date - timedelta(days=7),
                            'Column_Type': 'On hand',
                            'Value': float(oh),
                        })

                # Extract forecast values
                d0 = cols['date0']
                for off, ci in enumerate(range(d0, L)):
                    val = row[ci]
                    if isinstance(val, (int, float)):
                        records.append({
                            **base,
                            'Date': start_date + timedelta(days=off * 7),
                            'Column_Type': 'Forecast',
                            'Value': float(val),
                        })

            all_records.extend(records)

        return pd.DataFrame(all_records) if all_records else None

    except Exception:
        return None


def load_all_files(uploaded_files):
    """Load and combine all uploaded files with caching"""
    # Create fingerprint for cache invalidation
    fingerprint = hashlib.md5(
        b''.join(f.name.encode() + str(f.size).encode() for f in uploaded_files)
    ).hexdigest()

    # Return cached data if available
    if (st.session_state.get('_file_fp') == fingerprint
            and 'combined_df' in st.session_state):
        return st.session_state['combined_df'], st.session_state.get('_parse_status', [])

    # Parse all files
    all_data, status = [], []
    for f in uploaded_files:
        f.seek(0)
        fb = f.read()
        df = parse_one_file(fb, f.name)
        if df is not None and len(df):
            all_data.append(df)
            status.append((f.name, '✓ Success', len(df)))
        else:
            status.append((f.name, '✗ Failed', 0))

    # Combine and cache
    combined = pd.concat(all_data, ignore_index=True) if all_data else None
    st.session_state['combined_df'] = combined
    st.session_state['_file_fp'] = fingerprint
    st.session_state['_parse_status'] = status
    return combined, status


# ═══════════════════════════════════════════════════════════════
#  Filter Components
# ═══════════════════════════════════════════════════════════════

def make_filter(label: str, options: list, key: str, default_selection: list = None) -> list:
    """
    Multiselect with Select All / Clear buttons
    COMPLETELY FIXED: No more auto-reselection!

    Args:
        label: Filter label
        options: Available options
        key: Session state key
        default_selection: Custom default selections (None = select all)

    Key fix: Track what options existed LAST TIME to detect real changes
    """
    # Initialize if first time
    if key not in st.session_state:
        if default_selection is not None:
            # Use custom default (only items that exist in options)
            st.session_state[key] = [opt for opt in default_selection if opt in options]
        else:
            # Default: select all
            st.session_state[key] = list(options)

    # CRITICAL: Store what options were available in previous run
    # This lets us distinguish between:
    # 1. User deselected something (don't re-add it)
    # 2. New file uploaded (do add new options)
    prev_key = f"{key}__prev_options"

    if prev_key not in st.session_state:
        # First run - store current options
        st.session_state[prev_key] = set(options)

    # Get current state
    current_options = set(options)
    previous_options = st.session_state[prev_key]
    current_selections = set(st.session_state[key])

    # Detect if options actually changed (file added/removed)
    if current_options != previous_options:
        # Options changed! Update smartly

        # Find TRULY new options (from new files)
        truly_new_options = current_options - previous_options

        # Keep existing valid selections + add truly new ones
        valid_selections = current_selections & current_options
        st.session_state[key] = list(valid_selections | truly_new_options)

        # Update the tracker
        st.session_state[prev_key] = current_options
    else:
        # Options didn't change - just clean up invalid selections
        # (This handles case where user deselects something)
        st.session_state[key] = [v for v in st.session_state[key] if v in current_options]

    # Select All / Clear buttons
    col_a, col_b = st.columns(2)
    if col_a.button("Select All", key=f"_a_{key}", use_container_width=True):
        st.session_state[key] = list(options)
        st.rerun()
    if col_b.button("Clear", key=f"_b_{key}", use_container_width=True):
        st.session_state[key] = []
        st.rerun()

    return st.multiselect(label, options=options, key=key,
                          label_visibility="collapsed")


# ═══════════════════════════════════════════════════════════════
#  Visualization
# ═══════════════════════════════════════════════════════════════

def is_wos_related(text: str) -> bool:
    """Check if text contains 'wos' (case-insensitive)"""
    return 'wos' in text.lower() if text else False


def build_pivot_tables(desc_agg: pd.DataFrame, metric_label: str = ''):
    """
    Build TWO pivot tables:
    1. Weekly data table (sorted by WeekSortKey)
    2. Delta table (consecutive week differences)

    Now includes ISO week numbers as first row
    """
    oh_df = desc_agg[desc_agg['Column_Type'] == 'On hand']
    fct_df = desc_agg[desc_agg['Column_Type'] == 'Forecast']

    if fct_df.empty:
        return pd.DataFrame(), pd.DataFrame()

    # Get week information
    week_info = fct_df[['WeekSortKey', 'WeekLabel']].drop_duplicates()
    week_info = week_info.sort_values('WeekSortKey', ascending=False)

    # Create pivot table
    pt = fct_df.pivot_table(index='WeekSortKey', columns='Date',
                            values='Value', aggfunc='sum').T.sort_index()

    key_to_label = dict(zip(week_info['WeekSortKey'], week_info['WeekLabel']))

    # Build weekly table
    weekly_rows = []

    # FIRST ROW: ISO Week Numbers for each date
    iso_week_row = {'Metric': 'ISO Week'}
    iso_week_row['On hand'] = ''  # Empty for On hand column
    for dt in pt.index:
        date_str = dt.strftime('%Y-%m-%d')
        # Get ISO week number
        iso_week = dt.isocalendar()[1]  # Returns (year, week, weekday)
        iso_week_row[date_str] = f'W{iso_week}'
    weekly_rows.append(iso_week_row)

    # REMAINING ROWS: Actual data
    for sort_key in week_info['WeekSortKey']:
        label = key_to_label[sort_key]
        r = {'Metric': label}

        # Get On hand for this week
        oh_rows = oh_df[oh_df['WeekSortKey'] == sort_key]
        oh_val = oh_rows['Value'].sum() if not oh_rows.empty else None
        r['On hand'] = oh_val

        # Get forecast values for each date
        for dt in pt.index:
            if sort_key in pt.columns:
                v = pt.loc[dt, sort_key]
                date_str = dt.strftime('%Y-%m-%d')
                r[date_str] = v if pd.notna(v) else None
        weekly_rows.append(r)

    weekly_table = pd.DataFrame(weekly_rows)

    # Build delta table
    delta_rows = []
    sort_keys = list(week_info['WeekSortKey'])

    for i in range(len(sort_keys) - 1):
        newer_key = sort_keys[i]
        older_key = sort_keys[i + 1]
        newer_label = key_to_label[newer_key]
        older_label = key_to_label[older_key]

        r = {'Metric': f'{newer_label}→{older_label}'}

        # Delta for On hand
        newer_oh = oh_df[oh_df['WeekSortKey'] == newer_key]
        older_oh = oh_df[oh_df['WeekSortKey'] == older_key]

        if not newer_oh.empty and not older_oh.empty:
            r['On hand'] = newer_oh['Value'].sum() - older_oh['Value'].sum()
        else:
            r['On hand'] = None

        # Delta for each date
        for dt in pt.index:
            newer_val = pt.loc[dt, newer_key] if newer_key in pt.columns else None
            older_val = pt.loc[dt, older_key] if older_key in pt.columns else None
            date_str = dt.strftime('%Y-%m-%d')

            if pd.notna(newer_val) and pd.notna(older_val):
                r[date_str] = newer_val - older_val
            else:
                r[date_str] = None
        delta_rows.append(r)

    delta_table = pd.DataFrame(delta_rows)

    return weekly_table, delta_table


def style_table(df: pd.DataFrame, is_wos: bool = False):
    """Apply styling to table"""

    def apply_color(val):
        if isinstance(val, (int, float)) and val < 0:
            return 'color:#d62728;font-weight:bold'
        return ''

    styled = df.style.applymap(apply_color, subset=[c for c in df.columns if c != 'Metric'])
    numeric_cols = [c for c in df.columns if c != 'Metric']

    # Format function that handles both strings and numbers
    def safe_format(val, fmt):
        if pd.isna(val) or val == '':
            return ''
        if isinstance(val, str):
            return val  # Return strings as-is (for ISO Week row)
        try:
            return fmt.format(val)
        except (ValueError, TypeError):
            return str(val)

    # Format: 2 decimals for WOS, integers for others
    if is_wos:
        format_dict = {col: lambda x: safe_format(x, '{:.2f}') for col in numeric_cols}
    else:
        format_dict = {col: lambda x: safe_format(x, '{:.0f}') for col in numeric_cols}

    styled = styled.format(format_dict, na_rep='')
    return styled


# ═══════════════════════════════════════════════════════════════
#  Main Program
# ═══════════════════════════════════════════════════════════════

def main():
    with st.sidebar:
        st.markdown("## 📁 Upload Files")
        uploaded = st.file_uploader(
            "Select Excel files (format: MPA YYYYwkNN.xlsx)",
            type=['xlsx'],
            accept_multiple_files=True,
            label_visibility="collapsed",
        )

        if not uploaded:
            st.info("Please upload DDSS files\n\nFormat: `MPA YYYYwk00.xlsx`")
            st.session_state.pop('combined_df', None)
            return

        # SDOS file upload
        st.markdown("## 📊 SDOS File (Optional)")
        st.caption("Upload SDOS file to enable WOS calculation")
        sdos_file = st.file_uploader(
            "Upload SDOS.xlsx",
            type=['xlsx'],
            key='sdos_file',
        )

        # Parse DDSS files
        with st.spinner("Parsing files..."):
            combined, status = load_all_files(uploaded)

        if combined is None:
            st.error("Parse failed, please check file format")
            return

        # Check if files changed (show notification)
        current_files = set(f.name for f in uploaded)
        if '_last_files' in st.session_state:
            last_files = st.session_state['_last_files']
            new_files = current_files - last_files
            removed_files = last_files - current_files

            if new_files:
                st.success(f"✓ Added: {', '.join(new_files)}")
            if removed_files:
                st.info(f"ℹ️ Removed: {', '.join(removed_files)}")

        st.session_state['_last_files'] = current_files

        # Calculate WOS if SDOS provided
        if sdos_file:
            # Extract MPA from first filename
            first_filename = uploaded[0].name if uploaded else None
            mpa_from_file = None
            if first_filename:
                mpa_from_file, _, _ = extract_mpa_year_week(first_filename)

            with st.spinner("Loading SDOS and calculating WOS..."):
                sdos_file.seek(0)
                sdos_bytes = sdos_file.read()
                sdos_data = load_sdos_file(sdos_bytes, mpa_from_file)

                if sdos_data:
                    combined = calculate_wos_for_dataframe(combined, sdos_data)
                    st.session_state['combined_df'] = combined

        st.success(f"Loaded {len(uploaded)} files, {len(combined):,} records")
        st.divider()

        # ── Filters ──
        st.markdown("## 🔍 Filters")

        # ── Sheet Selection (Toggle Buttons) ──
        st.markdown("## 📋 Sheet Type")

        all_sheets = sorted(combined['Sheet'].dropna().unique())
        ddss_only = [s for s in all_sheets if s.strip().upper() == 'DDSS']
        without_only = [s for s in all_sheets if 'WITHOUT' in s.strip().upper() or "W'OUT" in s.strip().upper()]

        # Initialize sheet mode
        if 'sheet_mode' not in st.session_state:
            st.session_state['sheet_mode'] = 'DDSS'

        col1, col2 = st.columns(2)

        with col1:
            if st.button(
                    "📊 DDSS",
                    key="btn_ddss",
                    use_container_width=True,
                    type="primary" if st.session_state['sheet_mode'] == 'DDSS' else "secondary"
            ):
                st.session_state['sheet_mode'] = 'DDSS'
                st.rerun()

        with col2:
            if st.button(
                    "📉 Without Unconfirmed",
                    key="btn_without",
                    use_container_width=True,
                    type="primary" if st.session_state['sheet_mode'] == 'Without' else "secondary"
            ):
                st.session_state['sheet_mode'] = 'Without'
                st.rerun()

        # Apply sheet filter - STRICT matching
        sheet_mode = st.session_state['sheet_mode']
        if sheet_mode == 'DDSS':
            sel_sheet = ddss_only
        else:  # 'Without'
            sel_sheet = without_only

        # CRITICAL: If no matching sheets, show message and stop
        if not sel_sheet:
            if sheet_mode == 'Without':
                st.info("ℹ️ No 'Without Unconfirmed Orders' sheets found in uploaded files")
            else:
                st.warning("⚠️ No DDSS sheets found")
            return

        w = combined[combined['Sheet'].isin(sel_sheet)]
        st.divider()

        # MPA Filter
        st.markdown("**MPA**")
        all_mpas = sorted(w['MPA'].dropna().unique())
        sel_mpa = make_filter("MPA", all_mpas, "f_mpa")
        if not sel_mpa:
            st.warning("Please select MPA")
            return
        w = w[w['MPA'].isin(sel_mpa)]
        st.divider()

        # Type Filter
        all_types = sorted(w['Type'].dropna().unique())
        if all_types:
            st.markdown("**Type / Details**")
            sel_type = make_filter("Type", all_types, "f_type")
            if not sel_type:
                st.warning("Please select Type")
                return
            w = w[w['Type'].isin(sel_type)]
            st.divider()

        # Part Number Filter
        all_parts = sorted(w['Consign_PN'].dropna().unique())
        st.markdown("**Part Number**")
        sel_part = make_filter("Part", all_parts, "f_part")
        if not sel_part:
            st.warning("Please select Part")
            return
        w = w[w['Consign_PN'].isin(sel_part)]
        st.divider()

        # Data Description Filter
        all_descs = sorted(w['Data_Description'].dropna().unique())
        st.markdown("**Data Description**")

        # Default selections for Data Description
        default_descs = [
            'Balance',
            'POR demand',
            'AOS (cover next week demand)',
            'SupplierHP (Confirmed Orders)',
            'SupplierHP (Unconfirmed Orders)',
            'WOS'
        ]

        sel_desc = make_filter("Desc", all_descs, "f_desc", default_selection=default_descs)
        if not sel_desc:
            st.warning("Please select Description")
            return
        st.divider()

        # Month Filter
        w['YearMonth'] = pd.to_datetime(w['Date']).dt.to_period('M')
        all_months = sorted(w['YearMonth'].dropna().unique().astype(str))
        st.markdown("**Month**")
        sel_months = make_filter("Month", all_months, "f_month")
        if not sel_months:
            st.warning("Please select Month")
            return
        w = w[w['YearMonth'].astype(str).isin(sel_months)]

    # ── Main Area ──
    st.title("📊 DDSS Forecast Analysis")
    if sdos_file:
        st.caption("✨ WOS calculation enabled")

    if 'combined_df' not in st.session_state or st.session_state['combined_df'] is None:
        st.info("👈 Upload files to begin")
        return

    filtered = w[w['Data_Description'].isin(sel_desc)]
    if filtered.empty:
        st.warning("No data matches current filters")
        return

    # Summary metrics
    week_labels = sorted(combined['WeekLabel'].dropna().unique())
    wk_min = week_labels[0] if week_labels else "N/A"
    wk_max = week_labels[-1] if week_labels else "N/A"

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Files", len(uploaded))
    c2.metric("Week Range", f"{wk_min} – {wk_max}")
    c3.metric("MPAs", len(sel_mpa))
    c4.metric("Parts", len(sel_part))

    with st.expander("📄 Parse Status", expanded=False):
        st.dataframe(
            pd.DataFrame(status, columns=['Filename', 'Status', 'Records']),
            use_container_width=True,
            hide_index=True,
        )

    st.divider()

    # Build labels for charts
    mpa_label = " + ".join(sel_mpa)
    sheet_label = ", ".join(sel_sheet)
    part_label = ", ".join(sel_part) if len(sel_part) <= 3 else f"{len(sel_part)} Parts"

    download_tables = {}

    # Generate charts and tables for each data description
    for desc in sel_desc:
        is_wos = is_wos_related(desc)

        st.subheader(f"📈  {desc}")
        st.caption(f"MPA: **{mpa_label}** │ Sheet: **{sheet_label}** │ Part: {part_label}")

        sub = filtered[filtered['Data_Description'] == desc]
        agg = sub.groupby(['WeekSortKey', 'WeekLabel', 'Date', 'Column_Type'], as_index=False)['Value'].sum()

        # Line chart
        fct = agg[agg['Column_Type'] == 'Forecast']
        if not fct.empty:
            pp = (fct.pivot_table(index='Date', columns='WeekLabel',
                                  values='Value', aggfunc='sum')
                  .reset_index().sort_values('Date'))

            week_labels_chart = [c for c in pp.columns if c != 'Date']
            colors = px.colors.qualitative.Set2
            fig = go.Figure()

            # Hover template
            hover_template = '<b>%{fullData.name}</b><br>Date: %{x|%Y-%m-%d}<br>Value: %{y' + (
                ':.2f' if is_wos else ':,.0f') + '}<extra></extra>'

            # Add traces
            for i, wk_label in enumerate(week_labels_chart):
                tmp = pp[['Date', wk_label]].dropna()
                fig.add_trace(go.Scatter(
                    x=tmp['Date'], y=tmp[wk_label],
                    mode='lines+markers',
                    name=wk_label,
                    line=dict(width=2.5, color=colors[i % len(colors)]),
                    marker=dict(size=6),
                    hovertemplate=hover_template,
                ))

            # X-axis tick configuration
            actual_dates = sorted(pp['Date'].unique())
            if len(actual_dates) > 10:
                display_dates = actual_dates[::3]
            elif len(actual_dates) > 5:
                display_dates = actual_dates[::2]
            else:
                display_dates = actual_dates

            # Layout
            fig.update_layout(
                margin=dict(t=30, b=60, l=50, r=20), height=420,
                xaxis_title=None, yaxis_title="Value",
                hovermode='x unified',
                legend=dict(orientation="h", y=1.05, x=1,
                            xanchor="right", yanchor="bottom"),
                xaxis=dict(
                    tickformat='%Y-%m-%d',
                    tickmode='array',
                    tickvals=display_dates,
                    tickangle=-45,
                    tickfont=dict(size=10),
                ),
                yaxis=dict(separatethousands=True),
            )
            st.plotly_chart(fig, use_container_width=True)

        # Build tables
        weekly_table, delta_table = build_pivot_tables(agg, metric_label=desc)

        # Weekly table
        with st.expander(f"📋 {desc}", expanded=True):
            if weekly_table.empty:
                st.info("No data")
            else:
                h = min(420, (len(weekly_table) + 1) * 38 + 12)
                st.dataframe(
                    style_table(weekly_table, is_wos=is_wos),
                    use_container_width=True,
                    height=h,
                )
                download_tables[desc] = weekly_table

        # Delta table
        with st.expander(f"📋 {desc} Delta", expanded=True):
            if delta_table.empty:
                st.info("No delta")
            else:
                h = min(420, (len(delta_table) + 1) * 38 + 12)
                st.dataframe(
                    style_table(delta_table, is_wos=is_wos),
                    use_container_width=True,
                    height=h,
                )
                download_tables[f"{desc} Delta"] = delta_table

        st.divider()

    # Export section
    st.subheader("💾 Export")

    # Prepare export data: pivot format with cleaned columns
    export_df = filtered.copy()

    # Remove specified columns
    cols_to_remove = ['Week', 'Year', 'WeekSortKey', 'Sheet', 'Column_Type', 'YearMonth']
    export_df = export_df.drop(columns=[c for c in cols_to_remove if c in export_df.columns], errors='ignore')

    # Round WOS to 2 decimal places
    wos_mask = export_df['Data_Description'] == 'WOS'
    if wos_mask.any():
        export_df.loc[wos_mask, 'Value'] = export_df.loc[wos_mask, 'Value'].round(2)

    # Pivot: rows = (MPA, Type, Consign_PN, Data_Description), columns = Date
    pivot_df = export_df.pivot_table(
        index=['MPA', 'Type', 'Consign_PN', 'Data_Description'],
        columns='Date',
        values='Value',
        aggfunc='first'
    ).reset_index()

    # Format date columns as YYYY-MM-DD
    date_cols = [col for col in pivot_df.columns if isinstance(col, pd.Timestamp)]
    for col in date_cols:
        # Rename column to formatted date string
        pivot_df.rename(columns={col: col.strftime('%Y-%m-%d')}, inplace=True)

    # Convert to CSV
    csv_data = pivot_df.to_csv(index=False, encoding='utf-8-sig')

    st.download_button(
        "⬇️ Download Pivot Data (CSV)",
        data=csv_data,
        file_name=f"{'_'.join(sel_mpa)}_pivot_data.csv",
        mime="text/csv",
        use_container_width=True,
    )

    with st.expander("🔍 Raw Data"):
        st.dataframe(filtered, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
